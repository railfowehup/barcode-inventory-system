# -*- coding: utf-8 -*-
"""导出/备份/恢复 API"""
import os
import json
from datetime import datetime

from flask import request, jsonify, send_file
from . import app, db, EXPORT_DIR
from db import records as db_records
from db import backup as db_backup


@app.route('/api/export', methods=['GET'])
def api_export():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    records = db_records.get_all_for_export(db,
        user_id=request.args.get('user_id'),
        status=request.args.get('status'),
        date_from=request.args.get('date_from'),
        date_to=request.args.get('date_to')
    )

    wb = Workbook()
    ws = wb.active
    ws.title = '包裹记录'

    headers = ['ID', '条码', '状态', '目的地', '重量(kg)', '操作人',
               '分拣人设备', '收件人', '物流单号', '签收人', '异常类型',
               '备注', '入库时间', '分拣时间', '签收时间']
    ws.append(headers)

    header_fill = PatternFill(start_color='FF2196F3', end_color='FF2196F3', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFFFF', size=12)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for r in records:
        ws.append([
            r.get('id', ''),
            r.get('barcode', ''),
            r.get('status', ''),
            r.get('address', ''),
            r.get('weight', 0),
            r.get('user_name', ''),
            r.get('device_id', '') or '',
            r.get('recipient', '') or '',
            r.get('logistics_no', '') or '',
            r.get('signer', '') or '',
            r.get('exception_type', '') or '',
            r.get('note', ''),
            r.get('created_at', ''),
            r.get('sort_at', '') or '',
            r.get('sign_time', '') or '',
        ])

    # 设置列宽（用字母映射，支持超过26列）
    col_letters = [chr(65 + i) if i < 26 else chr(64 + i // 26) + chr(65 + i % 26) for i in range(15)]
    col_widths = [10, 30, 10, 15, 12, 15, 20, 15, 20, 15, 15, 20, 25, 25, 25]
    for i, width in enumerate(col_widths):
        ws.column_dimensions[col_letters[i]].width = width

    file_name = f'包裹记录_{int(datetime.now().timestamp() * 1000)}.xlsx'
    file_path = os.path.join(EXPORT_DIR, file_name)
    wb.save(file_path)

    return send_file(file_path, as_attachment=True, download_name=file_name)


@app.route('/api/backup', methods=['GET'])
def api_backup():
    backup = db_backup.backup(db)
    file_name = f'backup_{int(datetime.now().timestamp() * 1000)}.json'
    file_path = os.path.join(EXPORT_DIR, file_name)
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(backup, f, ensure_ascii=False, indent=2)
    return send_file(file_path, as_attachment=True, download_name=file_name)


@app.route('/api/restore', methods=['POST'])
def api_restore():
    data = request.get_json()
    if not data or not data.get('data'):
        return jsonify({'error': '备份数据格式错误'}), 400

    try:
        users_restored, records_restored = db_backup.restore(db, data['data'])
        return jsonify({
            'success': True,
            'users_restored': users_restored,
            'records_restored': records_restored
        })
    except Exception as e:
        return jsonify({'error': f'恢复失败: {str(e)}'}), 500
