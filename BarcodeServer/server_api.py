# -*- coding: utf-8 -*-
"""
📦 包裹管理系统 - Flask API 服务
完全兼容原 Node.js server.js 的所有接口
"""

import os
import sys
import json
import socket
import base64
import io
from datetime import datetime

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS

from database import Database, BASE_DIR, EXPORT_DIR

app = Flask(__name__)
CORS(app)
db = Database()

# 允许桌面 GUI 注册回调，手机端提交后立即通知桌面刷新
update_listeners = []

def register_update_listener(callback):
    if callback not in update_listeners:
        update_listeners.append(callback)


def notify_update(event_type='update', data=None):
    for callback in list(update_listeners):
        try:
            callback(event_type, data)
        except Exception:
            pass

PORT = int(os.environ.get('PORT', 3000))
HOST = '0.0.0.0'

# 确保导出目录存在
os.makedirs(EXPORT_DIR, exist_ok=True)


# ==================== 工具函数 ====================

def get_local_ip():
    """获取本机局域网IP地址"""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except:
        pass
    try:
        hostname = socket.gethostname()
        return socket.gethostbyname(hostname)
    except:
        return '127.0.0.1'


def now_str():
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')


# ==================== API 路由 ====================

# 1. 用户登录/注册
@app.route('/api/login', methods=['POST'])
def api_login():
    data = request.get_json()
    if not data or not data.get('name'):
        return jsonify({'error': '请输入姓名'}), 400
    user = db.login_or_register(data['name'], data.get('role'))
    return jsonify(user)


# 2. 扫码入库
@app.route('/api/scan', methods=['POST'])
def api_scan():
    data = request.get_json() or {}
    # 日志方便调试：打印请求来源和内容
    try:
        print(f"[API] /api/scan from {request.remote_addr} -> {data}")
    except Exception:
        pass

    barcode = (data.get('barcode') or '').strip()
    user_id = data.get('user_id')
    user_name = (data.get('user_name') or '').strip()

    if not barcode:
        return jsonify({'error': '缺少条码'}), 400

    # 如果没有 user_id，但提交了 user_name，则自动注册/登陆该用户
    if not user_id:
        if user_name:
            user = db.login_or_register(user_name)
            if isinstance(user, dict) and user.get('id'):
                user_id = user['id']
        else:
            return jsonify({'error': '缺少用户信息（user_id 或 user_name）'}), 400

    user = db.get_user(user_id)
    if not user:
        return jsonify({'error': '用户不存在'}), 404

    existing = db.find_record_by_barcode_and_status(barcode, '入库')
    if existing:
        return jsonify({'error': '该包裹已入库', 'record': existing}), 409

    record = db.add_record(
        barcode=barcode,
        user_id=user_id,
        address=data.get('address', ''),
        weight=data.get('weight', 0),
        note=data.get('note', '')
    )
    notify_update('scan', record)
    return jsonify({'success': True, 'record': record})


# 3. 分拣
@app.route('/api/sort', methods=['POST'])
def api_sort():
    data = request.get_json()
    if not data or not data.get('barcode') or not data.get('user_id'):
        return jsonify({'error': '缺少条码或用户信息'}), 400

    user = db.get_user(data['user_id'])
    if not user:
        return jsonify({'error': '用户不存在'}), 404

    record = db.find_record_by_barcode(data['barcode'].strip())
    if not record:
        return jsonify({'error': '该包裹未入库，请先入库'}), 404

    if record['status'] == '分拣':
        return jsonify({'error': '该包裹已被分拣', 'record': record}), 409

    updated = db.update_record_status(
        record['id'], '分拣',
        device_id=(data.get('device_id', '') or '').strip(),
        sort_at=now_str()
    )
    notify_update('sort', updated)
    return jsonify({'success': True, 'record': updated})


# 4. 出库
@app.route('/api/ship', methods=['POST'])
def api_ship():
    data = request.get_json()
    if not data or not data.get('barcode') or not data.get('user_id'):
        return jsonify({'error': '缺少条码或用户信息'}), 400

    record = db.find_record_by_barcode(data['barcode'].strip())
    if not record:
        return jsonify({'error': '该包裹未入库'}), 404

    if record['status'] == '出库':
        return jsonify({'error': '该包裹已出库', 'record': record}), 409

    updated = db.update_record_status(
        record['id'], '出库',
        logistics_no=(data.get('logistics_no', '') or '').strip(),
        recipient=(data.get('recipient', '') or '').strip(),
        sort_at=now_str()
    )
    notify_update('ship', updated)
    return jsonify({'success': True, 'record': updated})


# 5. 签收
@app.route('/api/sign', methods=['POST'])
def api_sign():
    data = request.get_json()
    if not data or not data.get('barcode') or not data.get('user_id'):
        return jsonify({'error': '缺少条码或用户信息'}), 400

    record = db.find_record_by_barcode(data['barcode'].strip())
    if not record:
        return jsonify({'error': '该包裹未入库'}), 404

    if record['status'] == '签收':
        return jsonify({'error': '该包裹已签收', 'record': record}), 409

    exception_type = (data.get('exception_type', '') or '').strip()
    new_status = '异常' if exception_type else '签收'
    updated = db.update_record_status(
        record['id'], new_status,
        signer=(data.get('signer', '') or '').strip(),
        sign_time=now_str(),
        exception_type=exception_type
    )
    notify_update('sign', updated)
    return jsonify({'success': True, 'record': updated})


# 6. 改地址
@app.route('/api/records/<int:record_id>/address', methods=['PUT'])
def api_update_address(record_id):
    data = request.get_json()
    if not data or not data.get('address'):
        return jsonify({'error': '缺少地址'}), 400

    result = db.update_record_address(
        record_id, data['address'],
        changed_by=data.get('user_name', '未知')
    )
    if result is None:
        return jsonify({'error': '记录不存在'}), 404

    updated, history = result
    notify_update('address_change', updated)
    return jsonify({'success': True, 'record': updated, 'address_history': history})


# 7. 编辑包裹（管理员）
@app.route('/api/records/<int:record_id>', methods=['PUT'])
def api_update_record(record_id):
    data = request.get_json()
    user = db.get_user(data.get('user_id'))
    if not user or user.get('role') != 'admin':
        return jsonify({'error': '仅管理员可编辑'}), 403

    record = db.get_record(record_id)
    if not record:
        return jsonify({'error': '记录不存在'}), 404

    kwargs = {}
    if 'address' in data:
        kwargs['address'] = data['address'].strip()
    if 'weight' in data:
        kwargs['weight'] = float(data['weight'])
    if 'note' in data:
        kwargs['note'] = data['note'].strip()

    updated = db.update_record(record_id, **kwargs) if kwargs else record
    notify_update('record_update', updated)
    return jsonify({'success': True, 'record': updated})


# 8. 删除包裹（管理员）
@app.route('/api/records/<int:record_id>', methods=['DELETE'])
def api_delete_record(record_id):
    data = request.get_json()
    user = db.get_user(data.get('user_id'))
    if not user or user.get('role') != 'admin':
        return jsonify({'error': '仅管理员可删除'}), 403

    if db.delete_record(record_id):
        notify_update('delete', {'record_id': record_id})
        return jsonify({'success': True, 'message': '已删除'})
    return jsonify({'error': '记录不存在'}), 404


# 9. 搜索记录
@app.route('/api/records/search', methods=['GET'])
def api_search_records():
    result = db.search_records(
        q=request.args.get('q'),
        status=request.args.get('status'),
        address=request.args.get('address'),
        date_from=request.args.get('date_from'),
        date_to=request.args.get('date_to'),
        page=request.args.get('page', 1),
        page_size=request.args.get('page_size', 50)
    )
    return jsonify({
        'total': result['total'],
        'page': int(request.args.get('page', 1)),
        'page_size': int(request.args.get('page_size', 50)),
        'records': result['records']
    })


# 10. 查询记录（分页）
@app.route('/api/records', methods=['GET'])
def api_get_records():
    result = db.get_records(
        user_id=request.args.get('user_id'),
        status=request.args.get('status'),
        date_from=request.args.get('date_from'),
        date_to=request.args.get('date_to'),
        page=request.args.get('page', 1),
        page_size=request.args.get('page_size', 20)
    )
    return jsonify({
        'total': result['total'],
        'page': int(request.args.get('page', 1)),
        'page_size': int(request.args.get('page_size', 20)),
        'records': result['records']
    })


# 11. 统计
@app.route('/api/stats', methods=['GET'])
def api_stats():
    stats = db.get_stats(user_id=request.args.get('user_id'))
    return jsonify(stats)


# 12. 导出 Excel
@app.route('/api/export', methods=['GET'])
def api_export():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    records = db.get_all_records_for_export(
        user_id=request.args.get('user_id'),
        status=request.args.get('status'),
        date_from=request.args.get('date_from'),
        date_to=request.args.get('date_to')
    )

    wb = Workbook()
    ws = wb.active
    ws.title = '包裹记录'

    headers = ['ID', '条码', '状态', '目的地', '重量(kg)', '操作人',
               '分拣人设备', '收件人', '物流单号', '签收人', '异常类型',
               '备注', '入库时间', '分拣时间', '签收时间']
    ws.append(headers)

    # 表头样式
    header_fill = PatternFill(start_color='FF2196F3', end_color='FF2196F3', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFFFF', size=12)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for r in records:
        ws.append([
            r.get('id', ''),
            r.get('barcode', ''),
            r.get('status', ''),
            r.get('address', ''),
            r.get('weight', 0),
            r.get('user_name', ''),
            r.get('device_id', '') or '',
            r.get('recipient', '') or '',
            r.get('logistics_no', '') or '',
            r.get('signer', '') or '',
            r.get('exception_type', '') or '',
            r.get('note', ''),
            r.get('created_at', ''),
            r.get('sort_at', '') or '',
            r.get('sign_time', '') or '',
        ])

    # 设置列宽
    col_widths = [10, 30, 10, 15, 12, 15, 20, 15, 20, 15, 15, 20, 25, 25, 25]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else 'A'].width = width

    file_name = f'包裹记录_{int(datetime.now().timestamp() * 1000)}.xlsx'
    file_path = os.path.join(EXPORT_DIR, file_name)
    wb.save(file_path)

    return send_file(file_path, as_attachment=True, download_name=file_name)


# 13. 获取所有用户
@app.route('/api/users', methods=['GET'])
def api_users():
    users = db.get_all_users()
    return jsonify(users)


# 14. 合并手机数据
@app.route('/api/merge', methods=['POST'])
def api_merge():
    data = request.get_json()
    if not data or not data.get('records'):
        return jsonify({'error': '缺少数据'}), 400

    result = db.merge_data(
        users_data=data.get('users'),
        records_data=data['records'],
        device_id=data.get('device_id', ''),
        user_name=data.get('user_name', '')
    )
    return jsonify({'success': True, **result})


# 15. 电脑推送到手机
@app.route('/api/pull', methods=['POST'])
def api_pull():
    data = request.get_json() or {}
    users, records = db.pull_data(since=data.get('since'))
    return jsonify({
        'success': True,
        'users': users,
        'records': records,
        'total_records': len(records)
    })


# 16. 备份
@app.route('/api/backup', methods=['GET'])
def api_backup():
    backup = db.backup_data()
    file_name = f'backup_{int(datetime.now().timestamp() * 1000)}.json'
    file_path = os.path.join(EXPORT_DIR, file_name)
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(backup, f, ensure_ascii=False, indent=2)
    return send_file(file_path, as_attachment=True, download_name=file_name)


# 17. 恢复
@app.route('/api/restore', methods=['POST'])
def api_restore():
    data = request.get_json()
    if not data or not data.get('data'):
        return jsonify({'error': '备份数据格式错误'}), 400

    try:
        users_restored, records_restored = db.restore_data(data['data'])
        return jsonify({
            'success': True,
            'users_restored': users_restored,
            'records_restored': records_restored
        })
    except Exception as e:
        return jsonify({'error': f'恢复失败: {str(e)}'}), 500


# 18. 同步日志
@app.route('/api/sync-logs', methods=['GET'])
def api_sync_logs():
    logs = db.get_sync_logs()
    return jsonify(logs)


# 19. 批量同步
@app.route('/api/scan/batch', methods=['POST'])
def api_scan_batch():
    data = request.get_json() or {}
    try:
        print(f"[API] /api/scan/batch from {request.remote_addr} -> records={len(data.get('records') or [])}")
    except Exception:
        pass

    records = data.get('records') or []
    user_id = data.get('user_id')
    user_name = (data.get('user_name') or '').strip()

    if not records:
        return jsonify({'error': '缺少数据'}), 400

    if not user_id:
        if user_name:
            user = db.login_or_register(user_name)
            if isinstance(user, dict) and user.get('id'):
                user_id = user['id']
        else:
            return jsonify({'error': '缺少用户信息（user_id 或 user_name）'}), 400

    user = db.get_user(user_id)
    if not user:
        return jsonify({'error': '用户不存在'}), 404

    inserted = 0
    for item in records:
        ts = None
        if item.get('timestamp'):
            try:
                ts = datetime.fromisoformat(item['timestamp'].replace('Z', '+00:00'))
                ts = ts.strftime('%Y-%m-%d %H:%M:%S')
            except:
                pass
        existing = db.find_record_by_barcode((item.get('barcode') or '').strip())
        if not existing:
            db.add_record(
                barcode=item.get('barcode', ''),
                user_id=user_id,
                note=item.get('note', ''),
                status='入库',
                created_at=ts
            )
            inserted += 1

    if inserted > 0:
        notify_update('scan_batch', {'inserted': inserted})
    return jsonify({'success': True, 'inserted': inserted})


# 20. 健康检查
@app.route('/api/health', methods=['GET'])
def api_health():
    return jsonify({'status': 'ok', 'time': datetime.now().isoformat()})


# 21. 二维码生成
@app.route('/api/qrcode', methods=['GET'])
def api_qrcode():
    try:
        import qrcode as qrcode_lib
        ip = get_local_ip()
        server_url = f'http://{ip}:{PORT}'
        img = qrcode_lib.make(server_url)
        buffer = io.BytesIO()
        img.save(buffer, format='PNG')
        img_base64 = base64.b64encode(buffer.getvalue()).decode()
        data_url = f'data:image/png;base64,{img_base64}'
        return jsonify({
            'success': True,
            'url': server_url,
            'qrcode': data_url
        })
    except Exception as e:
        return jsonify({'error': f'二维码生成失败: {str(e)}'}), 500


# 22. 设备心跳 - 手机端定期上报在线状态
@app.route('/api/device/heartbeat', methods=['POST'])
def api_device_heartbeat():
    try:
        data = request.get_json()
        if not data or not data.get('device_id'):
            return jsonify({'error': '缺少 device_id'}), 400
        db.record_heartbeat(
            device_id=data['device_id'],
            device_name=data.get('device_name', ''),
            ip_address=data.get('ip_address', ''),
            user_name=data.get('user_name', '')
        )
        return jsonify({'success': True, 'time': now_str()})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# 23. 获取在线设备列表
@app.route('/api/devices/online', methods=['GET'])
def api_devices_online():
    try:
        timeout = request.args.get('timeout', 30, type=int)
        devices = db.get_online_devices(timeout)
        return jsonify({
            'success': True,
            'online_count': len(devices),
            'devices': devices
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# 24. 获取所有设备历史
@app.route('/api/devices/history', methods=['GET'])
def api_devices_history():
    try:
        limit = request.args.get('limit', 50, type=int)
        devices = db.get_all_devices(limit)
        return jsonify({
            'success': True,
            'devices': devices
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== 增量同步 API ====================

# 25. 增量推送 - 手机推送到电脑
@app.route('/api/sync/push', methods=['POST'])
def api_sync_push():
    """手机推送增量记录到电脑"""
    try:
        data = request.get_json() or {}
        records = data.get('records', [])
        device_id = data.get('device_id', '')
        user_name = data.get('user_name', '')

        if not records:
            return jsonify({'success': True, 'inserted': 0, 'updated': 0, 'conflicts': 0})

        inserted = 0
        updated = 0
        conflicts = []

        for item in records:
            barcode = (item.get('barcode') or '').strip()
            if not barcode:
                continue

            existing = db.find_record_by_barcode(barcode)
            if existing:
                # 版本控制：手机版本 > 电脑版本 则覆盖
                phone_version = item.get('version', 1)
                server_version = existing.get('version', 1)
                if phone_version > server_version:
                    # 手机版本更新，覆盖
                    db.update_record_status(
                        existing['id'],
                        item.get('status', existing['status']),
                        address=item.get('address', existing.get('address', '')),
                        weight=item.get('weight', existing.get('weight', 0)),
                        note=item.get('note', existing.get('note', '')),
                        updated_at=item.get('updated_at', now_str()),
                        version=phone_version
                    )
                    updated += 1
                elif phone_version == server_version:
                    # 版本相同，标记冲突
                    conflicts.append({
                        'barcode': barcode,
                        'server_record': existing,
                        'phone_record': item
                    })
                # 手机版本 < 电脑版本，忽略
            else:
                # 新记录，直接插入
                db.add_record(
                    barcode=barcode,
                    user_id=item.get('user_id', 1),
                    address=item.get('address', ''),
                    weight=item.get('weight', 0),
                    note=item.get('note', ''),
                    status=item.get('status', '入库'),
                    created_at=item.get('created_at')
                )
                inserted += 1

        if inserted > 0 or updated > 0:
            notify_update('sync_push', {'inserted': inserted, 'updated': updated})

        return jsonify({
            'success': True,
            'inserted': inserted,
            'updated': updated,
            'conflicts': conflicts,
            'conflict_count': len(conflicts)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# 26. 增量拉取 - 手机从电脑拉取
@app.route('/api/sync/pull', methods=['POST'])
def api_sync_pull():
    """手机从电脑拉取增量记录"""
    try:
        data = request.get_json() or {}
        since = data.get('since')  # 上次同步时间
        device_id = data.get('device_id', '')

        users, records = db.pull_data(since=since)

        return jsonify({
            'success': True,
            'users': users,
            'records': records,
            'total_records': len(records),
            'server_time': now_str()
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# 27. 同步状态查询
@app.route('/api/sync/status', methods=['GET'])
def api_sync_status():
    """查询同步状态"""
    try:
        device_id = request.args.get('device_id', '')
        stats = db.get_stats()
        return jsonify({
            'success': True,
            'total_records': stats.get('total', 0),
            'server_time': now_str(),
            'device_id': device_id
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# 28. 设备分组管理
@app.route('/api/devices/group', methods=['PUT'])
def api_device_group():
    """更新设备分组"""
    try:
        data = request.get_json() or {}
        device_id = data.get('device_id')
        group = data.get('group', '')
        if not device_id:
            return jsonify({'error': '缺少 device_id'}), 400
        db.update_device_group(device_id, group)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# 29. 设备离线报警检查
@app.route('/api/devices/check-offline', methods=['GET'])
def api_check_offline():
    """检查是否有设备离线（超过60秒无心跳）"""
    try:
        timeout = request.args.get('timeout', 60, type=int)
        devices = db.get_online_devices(timeout)
        all_devices = db.get_all_devices(100)
        offline_devices = []
        for dev in all_devices:
            if not any(d['device_id'] == dev['device_id'] for d in devices):
                offline_devices.append(dev)
        return jsonify({
            'success': True,
            'online_count': len(devices),
            'offline_count': len(offline_devices),
            'offline_devices': offline_devices
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== 启动服务器 ====================

def start_server(host=HOST, port=PORT):
    """启动 Flask 服务器"""
    print('========================================')
    print('  📦 包裹入库管理系统 - Python 版')
    print('========================================')
    ip = get_local_ip()
    print(f'  本机 IP: {ip}')
    print(f'  端口: {port}')
    print(f'  🌐 API: http://{ip}:{port}')
    print('========================================')
    print('  API 接口:')
    print('  POST /api/login          - 用户登录')
    print('  POST /api/scan           - 扫码入库')
    print('  POST /api/sort           - 分拣包裹')
    print('  POST /api/ship           - 出库')
    print('  POST /api/sign           - 签收')
    print('  PUT  /api/records/:id    - 编辑包裹')
    print('  PUT  /api/records/:id/address - 改地址')
    print('  DELETE /api/records/:id  - 删除包裹')
    print('  GET  /api/records        - 查询记录')
    print('  GET  /api/records/search - 搜索记录')
    print('  GET  /api/stats          - 统计数据')
    print('  GET  /api/export         - 导出Excel')
    print('  GET  /api/users          - 用户列表')
    print('  GET  /api/backup         - 备份数据')
    print('  POST /api/restore        - 恢复数据')
    print('  POST /api/merge          - 合并手机数据')
    print('  POST /api/pull           - 电脑推送到手机')
    print('  GET  /api/sync-logs      - 同步日志')
    print('  GET  /api/health         - 健康检查')
    print('  GET  /api/qrcode         - 二维码')
    print('  POST /api/device/heartbeat - 设备心跳')
    print('  GET  /api/devices/online   - 在线设备')
    print('  GET  /api/devices/history  - 设备历史')
    print('  POST /api/sync/push      - 增量推送')
    print('  POST /api/sync/pull      - 增量拉取')
    print('  GET  /api/sync/status    - 同步状态')
    print('  PUT  /api/devices/group  - 设备分组')
    print('  GET  /api/devices/check-offline - 离线检查')
    print('========================================')
    app.run(host=host, port=port, debug=False, use_reloader=False)


if __name__ == '__main__':
    start_server()
